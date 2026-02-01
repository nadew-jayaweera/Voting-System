import os
from dotenv import load_dotenv
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file
from flask_socketio import SocketIO, emit
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import io

load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('FLASK_SECRET', 'fallback_secret')
socketio = SocketIO(app)

# --- GLOBAL STORAGE ---
current_state = {'mode': 'LANDING'} 
voted_voter_ids = set() 

def connect_db():
    conn = sqlite3.connect('voting.db')
    conn.row_factory = sqlite3.Row
    return conn

# --- ROUTES ---

@app.route('/')
def index():
    return render_template('welcome.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        password = request.form.get('password')
        if password == os.getenv('ADMIN_PASSWORD'):
            session['logged_in'] = True
            return redirect(os.getenv('ADMIN_URL_PATH'))
        else:
            error = "Invalid Password"
    return render_template('login.html', error=error)

# --- NEW LOGOUT ROUTE ---
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route(os.getenv('ADMIN_URL_PATH', '/admin'))
def admin():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    return render_template('admin.html')

@app.route('/vote')
def vote():
    return render_template('voter.html')

@app.route('/screen')
def screen():
    return render_template('screen.html')

@app.route('/export_votes')
def export_votes():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Voting Results"
    
    # Add title
    ws['A1'] = 'Voting System - Results Export'
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A1:D1')
    
    # Add export date
    ws['A2'] = f'Exported on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A2:D2')
    
    # Add headers
    headers = ['ID', 'Contestant Name', 'Yes Votes', 'No Votes']
    ws.append([])
    ws.append(headers)
    
    # Style headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4f46e5", end_color="4f46e5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Fetch data from database
    conn = connect_db()
    cur = conn.cursor()
    cur.execute("SELECT id, name, yes_votes, no_votes FROM contestants ORDER BY yes_votes DESC")
    rows = cur.fetchall()
    conn.close()
    
    # Add data rows
    for row in rows:
        ws.append([row['id'], row['name'], row['yes_votes'], row['no_votes']])
    
    # Add total row
    total_yes = sum(row['yes_votes'] for row in rows)
    total_no = sum(row['no_votes'] for row in rows)
    ws.append([])
    ws.append(['', 'TOTAL', total_yes, total_no])
    
    # Style total row
    last_row = ws.max_row
    for col in range(1, 5):
        cell = ws.cell(row=last_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="e0e7ff", end_color="e0e7ff", fill_type="solid")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Save to bytes buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with timestamp
    filename = f'voting_results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# --- SOCKET EVENTS ---

@socketio.on('connect')
def handle_connect():
    emit('update_screen', {'mode': current_state['mode']})
    if current_state['mode'] in ['VOTING', 'STOPPED', 'ENDED']:
        conn = connect_db()
        cur = conn.cursor()
        cur.execute("SELECT id, name, yes_votes FROM contestants")
        rows = cur.fetchall()
        scores = {row['id']: {'name': row['name'], 'votes': row['yes_votes']} for row in rows}
        conn.close()
        emit('update_scores', scores)

@socketio.on('register_voter')
def handle_register_voter(data):
    voter_id = data.get('voter_id')
    
    # Check if already voted
    if voter_id and voter_id in voted_voter_ids:
        emit('vote_confirmed', {'success': False, 'message': 'You have already voted.'})
        return

    if current_state['mode'] == 'LANDING':
        # Send waiting status when in LANDING mode
        emit('voting_status', {'status': 'waiting'})
    elif current_state['mode'] in ['VOTING', 'STOPPED', 'ENDED']:
        conn = connect_db()
        cur = conn.cursor()
        cur.execute("SELECT id, name, yes_votes FROM contestants")
        rows = cur.fetchall()
        scores = {row['id']: {'name': row['name'], 'votes': row['yes_votes']} for row in rows}
        emit('update_scores', scores)

        if current_state['mode'] == 'VOTING':
            cur.execute("SELECT id, name FROM contestants")
            contestants = [{'id': row['id'], 'name': row['name']} for row in cur.fetchall()]
            emit('voting_status', {'status': 'open', 'contestants': contestants})
        elif current_state['mode'] in ['STOPPED', 'ENDED']:
            emit('voting_status', {'status': 'closed'})
        conn.close()

@socketio.on('show_landing')
def handle_show_landing():
    current_state['mode'] = 'LANDING'
    emit('update_screen', {'mode': 'LANDING'}, broadcast=True)
    emit('voting_status', {'status': 'waiting'}, broadcast=True)

@socketio.on('open_voting_session')
def handle_open_voting():
    global voted_voter_ids
    voted_voter_ids.clear() 
    current_state['mode'] = 'VOTING'

    conn = connect_db()
    cur = conn.cursor()
    cur.execute("SELECT id, name FROM contestants")
    all_contestants = [{'id': row['id'], 'name': row['name']} for row in cur.fetchall()]
    cur.execute("SELECT id, name, yes_votes FROM contestants")
    rows = cur.fetchall()
    scores = {row['id']: {'name': row['name'], 'votes': row['yes_votes']} for row in rows}
    conn.close()
    
    emit('update_screen', {'mode': 'VOTING'}, broadcast=True)
    emit('voting_status', {'status': 'open', 'contestants': all_contestants}, broadcast=True)
    emit('update_scores', scores, broadcast=True)

@socketio.on('submit_votes')
def handle_vote_submission(data):
    voter_id = data.get('voter_id')
    if not voter_id:
        emit('vote_confirmed', {'success': False, 'message': '⚠️ Error: Invalid voter ID!'}, room=request.sid)
        return
    
    if voter_id in voted_voter_ids:
        emit('vote_confirmed', {'success': False, 'message': '⚠️ Error: You have already voted!'}, room=request.sid)
        return

    votes = data['votes']
    conn = connect_db()
    cur = conn.cursor()
    for c_id, vote_val in votes.items():
        if vote_val == 'yes':
            cur.execute("UPDATE contestants SET yes_votes = yes_votes + 1 WHERE id = ?", (c_id,))
        elif vote_val == 'no':
            cur.execute("UPDATE contestants SET no_votes = no_votes + 1 WHERE id = ?", (c_id,))
    conn.commit()
    
    cur.execute("SELECT id, name, yes_votes FROM contestants")
    rows = cur.fetchall()
    conn.close()

    scores = {}
    for row in rows:
        scores[row['id']] = {'name': row['name'], 'votes': row['yes_votes']}

    voted_voter_ids.add(voter_id)
    emit('update_scores', scores, broadcast=True)
    emit('vote_confirmed', {'success': True, 'message': 'Votes Submitted Successfully!'}, room=request.sid)

@socketio.on('stop_voting')
def handle_stop_voting():
    current_state['mode'] = 'STOPPED'
    emit('voting_status', {'status': 'closed'}, broadcast=True)
    emit('update_screen', {'mode': 'STOPPED'}, broadcast=True)
    
    # Send current scores to all clients
    conn = connect_db()
    cur = conn.cursor()
    cur.execute("SELECT id, name, yes_votes FROM contestants")
    rows = cur.fetchall()
    scores = {row['id']: {'name': row['name'], 'votes': row['yes_votes']} for row in rows}
    conn.close()
    emit('update_scores', scores, broadcast=True)

@socketio.on('end_competition')
def handle_end_competition():
    current_state['mode'] = 'ENDED'
    emit('update_screen', {'mode': 'ENDED'}, broadcast=True)

@socketio.on('admin_reset_data')
def handle_reset_data():
    global voted_voter_ids
    voted_voter_ids.clear() 
    
    conn = connect_db()
    cur = conn.cursor()
    cur.execute("UPDATE contestants SET yes_votes = 0, no_votes = 0")
    conn.commit()
    
    cur.execute("SELECT id, name, yes_votes FROM contestants")
    rows = cur.fetchall()
    scores = {row['id']: {'name': row['name'], 'votes': row['yes_votes']} for row in rows}
    conn.close()
    
    emit('update_scores', scores, broadcast=True)

if __name__ == '__main__':
    socketio.run(app, debug=True, host='0.0.0.0')